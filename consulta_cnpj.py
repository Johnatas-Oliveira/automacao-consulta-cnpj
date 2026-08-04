#!/usr/bin/env python3
"""
Consulta automática de CNPJs a partir de uma planilha (Excel ou CSV).

Convenção de pastas (criadas automaticamente do lado deste script):
    input/   -> coloque aqui a planilha com os CNPJs (qualquer nome de arquivo)
    output/  -> a planilha de resultado é salva aqui automaticamente

Fluxo:
    1. Localiza a planilha dentro de input/ (o nome do arquivo é ignorado).
    2. Limpa, valida e remove duplicados dos CNPJs da coluna informada.
    3. Consulta cada CNPJ na BrasilAPI (fonte principal) e, em caso de falha,
       tenta a ReceitaWS (fallback).
    4. Gera a planilha de resultado em output/.

Uso (o único parâmetro é o nome exato da coluna que contém os CNPJs):
    python consulta_cnpj.py "CNPJ"
    python consulta_cnpj.py "CNPJ da empresa"
"""

from __future__ import annotations

import argparse
import re
import sys
import time
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path

import httpx
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

# --------------------------------------------------------------------------- #
# Configuração
# --------------------------------------------------------------------------- #

BASE_DIR = Path(__file__).resolve().parent
INPUT_DIR = BASE_DIR / "input"
OUTPUT_DIR = BASE_DIR / "output"
EXTENSOES_ACEITAS = (".xlsx", ".xls", ".csv")

BRASILAPI_URL = "https://brasilapi.com.br/api/cnpj/v1/{cnpj}"
RECEITAWS_URL = "https://receitaws.com.br/v1/cnpj/{cnpj}"

MAX_TENTATIVAS = 3
TIMEOUT_SEGUNDOS = 15
PAUSA_ENTRE_CONSULTAS = 1.0  # segundos, para não estourar rate limit
PAUSA_RETRY_BASE = 2.0       # backoff exponencial entre tentativas

COLUNAS_SAIDA = [
    "CNPJ",
    "Razão Social",
    "Nome Fantasia",
    "Situação Cadastral",
    "Data da Situação",
    "Natureza Jurídica",
    "Porte",
    "Capital Social",
    "Data de Abertura",
    "CNAE Principal",
    "Logradouro",
    "Número",
    "Bairro",
    "Cidade",
    "UF",
    "CEP",
    "Telefone",
    "Email",
    "Sócios (QSA)",
    "Status",
]


# --------------------------------------------------------------------------- #
# Modelo de resultado
# --------------------------------------------------------------------------- #

@dataclass
class ResultadoConsulta:
    cnpj: str
    status: str = "Aguardando"  # Encontrado | Não encontrado | Erro
    dados: dict = field(default_factory=dict)


# --------------------------------------------------------------------------- #
# Localização de pastas e arquivo de entrada
# --------------------------------------------------------------------------- #

def preparar_pastas() -> None:
    """Garante que input/ e output/ existam do lado do script."""
    INPUT_DIR.mkdir(exist_ok=True)
    OUTPUT_DIR.mkdir(exist_ok=True)


def localizar_arquivo_entrada() -> Path:
    """Encontra a planilha dentro de input/ — o nome do arquivo é irrelevante,
    só o conteúdo importa. Exige que haja exatamente um arquivo válido."""
    candidatos = sorted(
        p for p in INPUT_DIR.iterdir()
        if p.is_file() and p.suffix.lower() in EXTENSOES_ACEITAS and not p.name.startswith(".")
    )

    if not candidatos:
        raise FileNotFoundError(
            f"Nenhuma planilha encontrada em '{INPUT_DIR}'. "
            f"Coloque um arquivo {'/'.join(EXTENSOES_ACEITAS)} nessa pasta e rode de novo."
        )

    if len(candidatos) > 1:
        nomes = "\n".join(f"  - {p.name}" for p in candidatos)
        raise FileExistsError(
            f"Mais de uma planilha encontrada em '{INPUT_DIR}':\n{nomes}\n"
            f"Deixe apenas UM arquivo nessa pasta e rode de novo."
        )

    return candidatos[0]


def localizar_coluna_cnpj(df: pd.DataFrame, coluna_informada: str) -> str:
    """Confirma que a coluna informada existe na planilha."""
    if coluna_informada not in df.columns:
        raise ValueError(
            f"Coluna '{coluna_informada}' não encontrada. "
            f"Colunas disponíveis: {list(df.columns)}"
        )
    return coluna_informada


def limpar_cnpj(valor) -> str | None:
    """Remove máscara e valida se restam exatamente 14 dígitos."""
    if pd.isna(valor):
        return None
    somente_digitos = re.sub(r"\D", "", str(valor))
    if len(somente_digitos) != 14:
        return None
    return somente_digitos


ENCODINGS_TENTATIVA = ["utf-8-sig", "utf-8", "cp1252", "latin1"]


def ler_csv_robusto(caminho: Path) -> pd.DataFrame:
    """Lê um CSV tentando detectar automaticamente o separador (',' ou ';')
    e o encoding (comum em exportações do Excel em pt-BR vir em cp1252/latin-1)."""
    ultimo_erro: Exception | None = None
    for encoding in ENCODINGS_TENTATIVA:
        try:
            # sep=None + engine="python" faz o pandas detectar o separador sozinho
            return pd.read_csv(caminho, dtype=str, sep=None, engine="python", encoding=encoding)
        except (UnicodeDecodeError, pd.errors.ParserError) as erro:
            ultimo_erro = erro
            continue
    raise ValueError(
        f"Não foi possível ler o CSV com nenhum dos encodings testados "
        f"({', '.join(ENCODINGS_TENTATIVA)}). Erro original: {ultimo_erro}"
    )


def carregar_cnpjs(caminho: Path, coluna_informada: str) -> list[str]:
    """Lê o arquivo de entrada (.xlsx, .xls ou .csv) e retorna uma lista de CNPJs únicos e válidos."""
    if caminho.suffix.lower() == ".csv":
        df = ler_csv_robusto(caminho)
    else:
        df = pd.read_excel(caminho, dtype=str)

    if df.empty:
        raise ValueError("Arquivo de entrada está vazio.")

    coluna = localizar_coluna_cnpj(df, coluna_informada)
    print(f"[INFO] Usando a coluna '{coluna}' como fonte dos CNPJs.")

    brutos = df[coluna].tolist()
    validos: list[str] = []
    invalidos = 0

    for valor in brutos:
        cnpj = limpar_cnpj(valor)
        if cnpj:
            validos.append(cnpj)
        elif not pd.isna(valor) and str(valor).strip():
            invalidos += 1

    total_antes_dedupe = len(validos)
    cnpjs_unicos = list(dict.fromkeys(validos))  # remove duplicados, preserva ordem
    duplicados = total_antes_dedupe - len(cnpjs_unicos)

    print(
        f"[INFO] {len(cnpjs_unicos)} CNPJs válidos encontrados "
        f"({invalidos} inválidos ignorados, {duplicados} duplicados removidos)."
    )
    return cnpjs_unicos


# --------------------------------------------------------------------------- #
# Consulta às APIs públicas
# --------------------------------------------------------------------------- #

def formatar_cnpj(cnpj: str) -> str:
    return f"{cnpj[0:2]}.{cnpj[2:5]}.{cnpj[5:8]}/{cnpj[8:12]}-{cnpj[12:14]}"


def consultar_brasilapi(client: httpx.Client, cnpj: str) -> dict | None:
    resp = client.get(BRASILAPI_URL.format(cnpj=cnpj), timeout=TIMEOUT_SEGUNDOS)
    if resp.status_code == 404:
        return None
    resp.raise_for_status()
    dados = resp.json()

    socios = dados.get("qsa") or []
    socios_fmt = "; ".join(
        f"{s.get('nome_socio', 'NÃO INFORMADO')} ({s.get('qualificacao_socio', 'NÃO INFORMADO')})"
        for s in socios
    ) or "NÃO INFORMADO"

    return {
        "Razão Social": dados.get("razao_social") or "NÃO INFORMADO",
        "Nome Fantasia": dados.get("nome_fantasia") or "NÃO INFORMADO",
        "Situação Cadastral": dados.get("descricao_situacao_cadastral") or "NÃO INFORMADO",
        "Data da Situação": dados.get("data_situacao_cadastral") or "NÃO INFORMADO",
        "Natureza Jurídica": dados.get("natureza_juridica") or "NÃO INFORMADO",
        "Porte": dados.get("porte") or "NÃO INFORMADO",
        "Capital Social": dados.get("capital_social") if dados.get("capital_social") is not None else "NÃO INFORMADO",
        "Data de Abertura": dados.get("data_inicio_atividade") or "NÃO INFORMADO",
        "CNAE Principal": dados.get("cnae_fiscal_descricao") or "NÃO INFORMADO",
        "Logradouro": dados.get("logradouro") or "NÃO INFORMADO",
        "Número": dados.get("numero") or "NÃO INFORMADO",
        "Bairro": dados.get("bairro") or "NÃO INFORMADO",
        "Cidade": dados.get("municipio") or "NÃO INFORMADO",
        "UF": dados.get("uf") or "NÃO INFORMADO",
        "CEP": dados.get("cep") or "NÃO INFORMADO",
        "Telefone": dados.get("ddd_telefone_1") or "NÃO INFORMADO",
        "Email": dados.get("email") or "NÃO INFORMADO",
        "Sócios (QSA)": socios_fmt,
    }


def consultar_receitaws(client: httpx.Client, cnpj: str) -> dict | None:
    resp = client.get(RECEITAWS_URL.format(cnpj=cnpj), timeout=TIMEOUT_SEGUNDOS)
    resp.raise_for_status()
    dados = resp.json()

    if dados.get("status") == "ERROR":
        return None

    socios = dados.get("qsa") or []
    socios_fmt = "; ".join(
        f"{s.get('nome', 'NÃO INFORMADO')} ({s.get('qual', 'NÃO INFORMADO')})"
        for s in socios
    ) or "NÃO INFORMADO"

    return {
        "Razão Social": dados.get("nome") or "NÃO INFORMADO",
        "Nome Fantasia": dados.get("fantasia") or "NÃO INFORMADO",
        "Situação Cadastral": dados.get("situacao") or "NÃO INFORMADO",
        "Data da Situação": dados.get("data_situacao") or "NÃO INFORMADO",
        "Natureza Jurídica": dados.get("natureza_juridica") or "NÃO INFORMADO",
        "Porte": dados.get("porte") or "NÃO INFORMADO",
        "Capital Social": dados.get("capital_social") or "NÃO INFORMADO",
        "Data de Abertura": dados.get("abertura") or "NÃO INFORMADO",
        "CNAE Principal": (dados.get("atividade_principal") or [{}])[0].get("text", "NÃO INFORMADO"),
        "Logradouro": dados.get("logradouro") or "NÃO INFORMADO",
        "Número": dados.get("numero") or "NÃO INFORMADO",
        "Bairro": dados.get("bairro") or "NÃO INFORMADO",
        "Cidade": dados.get("municipio") or "NÃO INFORMADO",
        "UF": dados.get("uf") or "NÃO INFORMADO",
        "CEP": dados.get("cep") or "NÃO INFORMADO",
        "Telefone": dados.get("telefone") or "NÃO INFORMADO",
        "Email": dados.get("email") or "NÃO INFORMADO",
        "Sócios (QSA)": socios_fmt,
    }


CAMPOS_COMPLEMENTAVEIS = ["Email", "Telefone"]


def completar_campos_faltantes(
    client: httpx.Client, cnpj: str, dados: dict, indice: int, total: int, cnpj_fmt: str
) -> dict:
    """Se algum campo em CAMPOS_COMPLEMENTAVEIS veio vazio da BrasilAPI, tenta
    preenchê-lo consultando a ReceitaWS (sem sobrescrever o que já foi encontrado)."""
    faltantes = [c for c in CAMPOS_COMPLEMENTAVEIS if dados.get(c) in (None, "NÃO INFORMADO")]
    if not faltantes:
        return dados

    for tentativa in range(1, MAX_TENTATIVAS + 1):
        try:
            extra = consultar_receitaws(client, cnpj)
            if extra:
                completados = []
                for campo in faltantes:
                    valor = extra.get(campo)
                    if valor and valor != "NÃO INFORMADO":
                        dados[campo] = valor
                        completados.append(campo)
                if completados:
                    print(
                        f"[{indice}/{total}] {cnpj_fmt}... complementado via ReceitaWS: "
                        f"{', '.join(completados)}"
                    )
            return dados
        except (httpx.TimeoutException, httpx.HTTPStatusError, httpx.TransportError):
            if tentativa < MAX_TENTATIVAS:
                time.sleep(PAUSA_RETRY_BASE * tentativa)
                continue
    return dados


def consultar_com_retry(client: httpx.Client, cnpj: str, indice: int, total: int) -> ResultadoConsulta:
    resultado = ResultadoConsulta(cnpj=cnpj)
    cnpj_fmt = formatar_cnpj(cnpj)

    for tentativa in range(1, MAX_TENTATIVAS + 1):
        try:
            dados = consultar_brasilapi(client, cnpj)
            if dados is not None:
                dados = completar_campos_faltantes(client, cnpj, dados, indice, total, cnpj_fmt)
                resultado.status = "Encontrado"
                resultado.dados = dados
                print(f"[{indice}/{total}] {cnpj_fmt}... OK (BrasilAPI)")
                return resultado
            break  # 404 confirmado — não adianta repetir na mesma fonte
        except (httpx.TimeoutException, httpx.HTTPStatusError, httpx.TransportError):
            if tentativa < MAX_TENTATIVAS:
                time.sleep(PAUSA_RETRY_BASE * tentativa)
                continue

    # Fallback: ReceitaWS
    for tentativa in range(1, MAX_TENTATIVAS + 1):
        try:
            dados = consultar_receitaws(client, cnpj)
            if dados is not None:
                resultado.status = "Encontrado"
                resultado.dados = dados
                print(f"[{indice}/{total}] {cnpj_fmt}... OK (ReceitaWS - fallback)")
                return resultado
            resultado.status = "Não encontrado"
            print(f"[{indice}/{total}] {cnpj_fmt}... NÃO ENCONTRADO")
            return resultado
        except (httpx.TimeoutException, httpx.HTTPStatusError, httpx.TransportError):
            if tentativa < MAX_TENTATIVAS:
                time.sleep(PAUSA_RETRY_BASE * tentativa)
                continue

    resultado.status = "Erro"
    print(f"[{indice}/{total}] {cnpj_fmt}... ERRO (ambas as fontes falharam)")
    return resultado


# --------------------------------------------------------------------------- #
# Geração da planilha de saída
# --------------------------------------------------------------------------- #

def gerar_planilha(resultados: list[ResultadoConsulta], caminho_saida: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Resultado CNPJ"

    # Cabeçalho
    ws.append(COLUNAS_SAIDA)
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    for cel in ws[1]:
        cel.fill = header_fill
        cel.font = header_font
        cel.alignment = Alignment(horizontal="center", vertical="center")

    # Linhas de dados
    for r in resultados:
        linha = [formatar_cnpj(r.cnpj)]
        for col in COLUNAS_SAIDA[1:-1]:  # entre CNPJ e Status
            linha.append(r.dados.get(col, "NÃO INFORMADO") if r.status == "Encontrado" else "-")
        linha.append(r.status)
        ws.append(linha)

    # Congela primeira linha e ativa autofiltro
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    # Ajuste automático (aproximado) da largura das colunas
    for col_idx, _ in enumerate(COLUNAS_SAIDA, start=1):
        letra = get_column_letter(col_idx)
        maior = max(
            (len(str(cel.value)) for cel in ws[letra] if cel.value is not None),
            default=10,
        )
        ws.column_dimensions[letra].width = min(max(maior + 2, 12), 45)

    wb.save(caminho_saida)


# --------------------------------------------------------------------------- #
# Programa principal
# --------------------------------------------------------------------------- #

def main() -> None:
    parser = argparse.ArgumentParser(
        description="Consulta automática de CNPJs com geração de planilha. "
        "Coloque a planilha de entrada na pasta input/ antes de rodar."
    )
    parser.add_argument("coluna", type=str, help="Nome EXATO da coluna que contém os CNPJs na planilha.")
    args = parser.parse_args()

    preparar_pastas()

    try:
        arquivo_entrada = localizar_arquivo_entrada()
    except (FileNotFoundError, FileExistsError) as erro:
        print(f"[ERRO] {erro}", file=sys.stderr)
        sys.exit(1)

    print(f"[INFO] Planilha de entrada: {arquivo_entrada.name}")

    inicio = time.time()
    try:
        cnpjs = carregar_cnpjs(arquivo_entrada, args.coluna)
    except ValueError as erro:
        print(f"[ERRO] {erro}", file=sys.stderr)
        sys.exit(1)

    if not cnpjs:
        print("[ERRO] Nenhum CNPJ válido foi encontrado no arquivo.", file=sys.stderr)
        sys.exit(1)

    resultados: list[ResultadoConsulta] = []
    total = len(cnpjs)
    interrompido = False

    with httpx.Client(headers={"User-Agent": "consulta-cnpj-script/1.0"}) as client:
        try:
            for i, cnpj in enumerate(cnpjs, start=1):
                resultado = consultar_com_retry(client, cnpj, i, total)
                resultados.append(resultado)
                if i < total:
                    time.sleep(PAUSA_ENTRE_CONSULTAS)
        except KeyboardInterrupt:
            interrompido = True
            print(
                f"\n[AVISO] Execução interrompida pelo usuário. "
                f"{len(resultados)}/{total} CNPJs já haviam sido consultados. "
                f"Gerando planilha com os resultados parciais..."
            )

    timestamp = datetime.now().strftime("%Y-%m-%d_%H%M")
    sufixo = "_PARCIAL" if interrompido else ""
    caminho_saida = OUTPUT_DIR / f"resultado_cnpj_{timestamp}{sufixo}.xlsx"

    if resultados:
        gerar_planilha(resultados, caminho_saida)
    else:
        print("[AVISO] Nenhum CNPJ foi consultado antes da interrupção — nenhuma planilha gerada.")
        sys.exit(1)

    encontrados = sum(1 for r in resultados if r.status == "Encontrado")
    nao_encontrados = sum(1 for r in resultados if r.status == "Não encontrado")
    erros = sum(1 for r in resultados if r.status == "Erro")
    duracao_min = (time.time() - inicio) / 60

    print("\n" + "=" * 50)
    print("CONSULTA INTERROMPIDA (RESULTADO PARCIAL)" if interrompido else "CONSULTA FINALIZADA")
    print(f"Consultados:      {len(resultados)}/{total}")
    print(f"Encontrados:      {encontrados}")
    print(f"Não encontrados:  {nao_encontrados}")
    print(f"Erros:            {erros}")
    print(f"Tempo total:      {duracao_min:.1f} min")
    print(f"Planilha gerada:  {caminho_saida}")
    print("=" * 50)

    if interrompido:
        sys.exit(130)  # código de saída convencional para interrupção via Ctrl+C


if __name__ == "__main__":
    main()
